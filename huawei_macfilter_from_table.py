from scrapli import Scrapli
from scrapli.exceptions import ScrapliException
import re
from openpyxl import load_workbook
import json


with open("config.json", "r") as config_file:
    config_data = json.load(config_file)


username = config_data["username"]
password = config_data["password"]


def get_mac_filter_entries(connection):

    try:
        result_mac_filter = connection.send_command("display security mac-filter")
        
        result_text = str(result_mac_filter.result)     
        
        # Splitting text into lines
        lines = result_text.split("\n")     
        
        # Initializing a list to store MAC addresses
        mac_addresses = []
        
        # Iterating through text lines
        for line in lines:
            mac_address_match = re.search(r"([0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4})", line)
            if mac_address_match:
                mac_address = mac_address_match.group(0)
                mac_addresses.append(mac_address)
        return mac_addresses
    except ScrapliException as e:
        print(f"An error occurred while getting MAC filter entries: {e}")
        return []


def read_excel_data(file_path):

    # Initialize a dictionary to store IP addresses and their corresponding MAC addresses
    ip_mac_data = {}

    try:

        # Load the Excel file
        wb = load_workbook(file_path, data_only=True)

        # Get the active worksheet
        ws = wb.active

        # Iterate through the rows in the worksheet, starting from the second row
        for row in ws.iter_rows(min_row=2, values_only=True):

            # Get the IP address from the 7th column
            ip_address = row[6]

            # Check if the IP address is not None and is in the correct format
            if ip_address is not None and re.match(r'^(\d{1,3}\.){3}\d{1,3}$', ip_address):
            
                # Initialize a list to store MAC addresses
                mac_addresses = []

                # Iterate through all values in the row, starting from the first column
                for value in row:
                    # Check if the value is a MAC address using regular expression
                    mac_address_match = re.search(r'([0-9A-Fa-f]{4}[\.-][0-9A-Fa-f]{4}[\.-][0-9A-Fa-f]{4})', str(value))
                    if mac_address_match:
                        mac_address = mac_address_match.group(0)
                        mac_addresses.append(mac_address)

                # Add the IP address and its corresponding MAC addresses to the dictionary
                ip_mac_data[ip_address] = mac_addresses
        
    except Exception as e:
        print(f"An error occurred while reading Excel data: {e}")

    finally:
        # Close the Excel file
        wb.close()

    return ip_mac_data


def update_mac_filter_entries(ip_mac_data, conn):
    try:
        for ip_address, mac_addresses in ip_mac_data.items():

            # Connect to the device
            conn.open()

            # Configuration mod
            config_mode_result = conn.send_command("config")

            # Getting current MAC filter entries
            existing_mac_addresses = get_mac_filter_entries(conn)
            
            # Deleting existing MAC addresses from MAC filter table
            for mac_address in existing_mac_addresses:
                delete_existing_mac_address = conn.send_command(f"undo security mac-filter source {mac_address}") 

            # Adding new MAC addresses from the dictionary
            for mac_address in mac_addresses:
                add_mac_address = conn.send_command(f"security mac-filter source {mac_address}") 

            # Getting updated MAC filter entries
            updated_mac_addresses = get_mac_filter_entries(conn)
            print(f"Updated MAC filter entries for IP address {ip_address}:")
            for mac_address in updated_mac_addresses:
                print(mac_address)

            # Saving the configuration
            save_config_result = conn.send_command("save")
            
            # Closing the connection
            conn.close()
            
    except ScrapliException as e:
        print(f"An error occurred: {e}")


# Load IP addresses and MAC addresses from Excel file
ip_mac_data = read_excel_data('macfilter_py.xlsx')


# Set up connection parameters
my_device = {
    "auth_username": username,
    "auth_password": password,
    "auth_strict_key": False,
    "platform": "huawei_smartax",
    "ssh_config_file": "my_ssh_config"  # Add this line to specify the path to the ssh_config file
                                        # my_ssh_config file must be: 
                                        #                                Host *
                                        #                                    HostKeyAlgorithms ssh-rsa
}


# Iterate through IP addresses and update MAC filter entries
for ip_address, mac_addresses in ip_mac_data.items():
    my_device["host"] = ip_address
    conn = Scrapli(**my_device)     # Create Scrapli object
    update_mac_filter_entries({ip_address: mac_addresses}, conn)



